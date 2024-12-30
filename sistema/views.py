from django.shortcuts import render,redirect,get_object_or_404
from django.urls import path,include
from django.contrib.auth import authenticate, login
from .models import *
from django.contrib import messages

from .forms import UsuarioForm, LoginForm,ParqueaderoForm,EditarUsuarioForm

from sistema import views    

# Create your views here.
def home(request):
    return render(request ,'index.html')
def error(request):
    return render(request ,'error.html')

def crear_usuario(request):
    if request.method == "POST":
        form = UsuarioForm(request.POST)
        if form.is_valid():
            form.save()  # Guarda el usuario con la contraseña encriptada
            return redirect('/sistema/users/tipouser/')  # Redirige a la vista deseada
    else:
        form = UsuarioForm()
    return render(request, 'crear_usuario.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            contrasena = form.cleaned_data['contrasena']
            usuario = authenticate(request, username=email, password=contrasena)
            if usuario is not None:
                login(request, usuario)
                
                # Redirige según el rol del usuario
                if usuario.rol.strip() == 'Administrador':
                    return redirect('/sistema/users/tipouser/')  # Cambia 'vista_administrador' por la URL o nombre de vista de administrador
                elif usuario.rol.strip() == 'Guarda':
                    return redirect('/sistema/parqueadero/')  # Cambia 'vista_guarda' por la URL o nombre de vista para guardas

                # En caso de un rol desconocido, puedes redirigir a una página de error o principal
                return redirect('pagina_principal')  # Opcional: redirigir a una página por defecto
            else:
                form.add_error(None, 'Email o contraseña incorrectos.')
    else:
        form = LoginForm()
    
    return render(request, 'login.html', {'form': form})



# --tipo_user

def irtipo_user(request):
    # Verifica si el usuario está autenticado
    if request.user.is_authenticated:
        try:
            # Filtra por el correo electrónico del usuario autenticado
            usuario_obj = Usuario.objects.get(email=request.user.email)
            
            # Verifica el rol del usuario
            if usuario_obj.rol.strip()  == 'Administrador':
                # Si es administrador, obtiene todos los usuarios y los pasa al template
                dato = Usuario.objects.filter(is_active=True)
                return render(request, 'administra/tipouser/Tipouser.html', {"dato": dato})
            else:
                # Si el usuario no tiene el rol adecuado, muestra un mensaje de error
                messages.error(request, "No tienes permiso para acceder a esta página.")
                return render(request, 'error.html', {"message": "Usted no tiene permisos para estar acá."})
        
        except Usuario.DoesNotExist:
            # Si no se encuentra un usuario con el correo dado, redirige a la página de login
            messages.error(request, "Usuario no encontrado.")
            return redirect("/sistema/login/")
    
    else:
        # Si el usuario no está autenticado, redirige al login
        return redirect("/sistema/login/")

    
    
def registartipo_user(request):
    if request.user.is_authenticated:
        empleado = Usuario.objects.get(email=request.user.email)
        if empleado.rol.strip() == 'Administrador':
            
            tipo_E=request.POST["tipo"]
            
        
            servio =Usuario.objects.create(
                tipo_empleado=tipo_E)
            
            return redirect("/administrador/users/tipouser/")
            
        else:
            # Si el usuario no tiene el tipo de usuario adecuado, mostrar un mensaje de alerta
            messages.error(request, "No tienes permiso para acceder a esta página.")
            return render(request, 'error.html', {"message": "usted no tiene permisos para esatr aca ."})    
    else:
        return redirect("/formularios/login/")
    

def eliminartipo_user(request, tipo):
    # Verifica si el usuario está autenticado
    if request.user.is_authenticated:
        try:
            # Obtiene el empleado autenticado
            empleado = Usuario.objects.get(email=request.user.email)
        
            # Verifica el rol del empleado
            if empleado.rol.strip() == 'Administrador':
                # Obtiene el tipo de empleado a desactivar
                tipo_E = Usuario.objects.get(email=tipo)
                
                # Cambia el estado a inactivo
                tipo_E.is_active = False
                tipo_E.save()  # Guarda los cambios en la base de datos
                
                # Redirige a la lista de tipos de usuarios
                messages.success(request, "El tipo de empleado ha sido desactivado.")
                return redirect("/sistema/users/tipouser/")
            else:
                # Si el usuario no tiene el rol adecuado, muestra un mensaje de error
                messages.error(request, "No tienes permiso para acceder a esta página.")
                return render(request, 'error.html', {"message": "Usted no tiene permisos para estar acá."})
        
        except Usuario.DoesNotExist:
            messages.error(request, "Usuario no encontrado.")
            return redirect("/administrador/users/tipouser/")
        except Exception as e:
            messages.error(request, f"Ocurrió un error: {str(e)}")
            return redirect("/administrador/users/tipouser/")
    else:
        return redirect("/formularios/login/")



def editar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    if request.method == 'POST':
        form = EditarUsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario actualizado correctamente.')
            return redirect('/sistema/users/tipouser/')  # Cambia esto por la vista donde quieres redirigir después de la edición
    else:
        form = EditarUsuarioForm(instance=usuario)
    return render(request, 'editar_usuario.html', {'form': form})

# --parqueadero

def listar_parqueaderos(request):
    if request.user.is_authenticated:
        try:
            usuario_obj = Usuario.objects.get(email=request.user.email)
            if usuario_obj.rol.strip() in ['Administrador', 'Guarda']:
                # Obtiene todos los parqueaderos activos
                parqueaderos = Parqueadero.objects.filter(estado='Activo')
                return render(request, 'administra/parqueadero/parqueadero.html', {'parqueaderos': parqueaderos})
            else:
                messages.error(request, "No tienes permiso para acceder a esta página.")
                return render(request, 'error.html', {"message": "Usted no tiene permisos para estar acá."})
        
        except Usuario.DoesNotExist:
            messages.error(request, "Usuario no encontrado.")
            return redirect("/sistema/login/")
    else:
        return redirect("/sistema/login/")


    

def registrar_parqueadero(request):
    if request.user.is_authenticated:
        try:
            usuario_obj = Usuario.objects.get(email=request.user.email)
            
            if usuario_obj.rol.strip() in ['Administrador', 'Guarda']:  # Verifica el rol correctamente
                if request.method == "POST":
                    form = ParqueaderoForm(request.POST)
                    if form.is_valid():
                        parqueadero = form.save(commit=False)  # No guarda inmediatamente
                        parqueadero.fecha_ingreso = timezone.now()  # Establece la fecha de ingreso actual
                        parqueadero.id_guarda = usuario_obj 
                        parqueadero.estado = "Activo" 

                        # Asignamos el conjunto 'zapan' a 'id_conjunto'
                        conjunto_obj = Conjunto.objects.get(nombre="zapan")
                        parqueadero.id_conjunto = conjunto_obj

                        parqueadero.save()  # Ahora guarda el objeto con la fecha de ingreso
                        return redirect('/sistema/parqueadero/')
                else:
                    form = ParqueaderoForm()  # Crea un nuevo formulario si no es POST
                
                return render(request, 'administra/parqueadero/registrar_parqueadero.html', {'form': form})
            else:
                messages.error(request, "No tienes permiso para acceder a esta página.")
                return render(request, 'error.html', {"message": "Usted no tiene permisos para estar acá."})
        
        except Usuario.DoesNotExist:
            messages.error(request, "Usuario no encontrado.")
            return redirect("/sistema/login/")
    else:
        return redirect("/sistema/login/")


    
import math
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from .models import Parqueadero, Conjunto

def salida_parqueadero(request, parqueadero_id):
    if request.user.is_authenticated:
        try:
            parqueadero = Parqueadero.objects.get(id=parqueadero_id)
            
            if parqueadero.estado == 'Activo':
                fecha_salida = timezone.now()
                horas_estacionado = (fecha_salida - parqueadero.fecha_ingreso).total_seconds() / 3600
                # Redondear hacia arriba siempre
                horas_estacionado_entero = math.ceil(horas_estacionado)

                conjunto = Conjunto.objects.get(id=parqueadero.id_conjunto.id)
                valor_fraccion_hora = Decimal(conjunto.fraccion_hora)

                valor_total = Decimal(horas_estacionado_entero) * valor_fraccion_hora
                valor_total_formateado = f"${valor_total:,.2f}"

                # Actualiza el estado del parqueadero

                # Muestra el mensaje
                messages.success(request, f"""El vehículo {parqueadero.placa_vehiculo} ha salido.<br>
                                              Tiempo estacionado: {horas_estacionado_entero} horas.<br>
                                              Total calculado: {valor_total_formateado}.<br>
                                              id :{parqueadero.id}

                                          """,extra_tags=str(parqueadero.id))
                return redirect('/sistema/parqueadero/')

            else:
                messages.error(request, "El parqueadero ya está finalizado.")
                return redirect("/sistema/parqueadero/")

        except Parqueadero.DoesNotExist:
            messages.error(request, "Parqueadero no encontrado.")
            return redirect("/sistema/parqueadero/")

    else:
        return redirect("/formularios/login/")






import os
from twilio.rest import Client
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from decimal import Decimal
from .models import Parqueadero, Conjunto
from django.core.mail import EmailMessage
from django.conf import settings
from django.shortcuts import redirect
from django.utils import timezone
from decimal import Decimal
from django.template.loader import render_to_string

from xhtml2pdf import pisa
from io import BytesIO
from django.template.loader import render_to_string

def generar_pdf_xhtml2pdf(parqueadero, valor_total, horas_estacionado, valor_fraccion_hora,fecha_salida):
    # Renderizar el HTML con el contexto de los datos
    html = render_to_string('factura_parqueadero.html', {
        'placa_vehiculo': parqueadero.placa_vehiculo,
        'fecha_ingreso': parqueadero.fecha_ingreso,
        'fecha_salida': fecha_salida,
        'horas_estacionado': horas_estacionado,
        'valor_fraccion_hora': valor_fraccion_hora,
        'valor_total': valor_total
    })
    
    # Crear un buffer de memoria para el PDF
    pdf_buffer = BytesIO()

    # Convertir el HTML a PDF usando xhtml2pdf
    pisa_status = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), pdf_buffer)
    
    # Volver al principio del buffer
    pdf_buffer.seek(0)

    # Si hay un error al generar el PDF, puedes manejarlo de la siguiente manera:
    if pisa_status.err:
        return None  # o puedes manejar el error como desees
    
    return pdf_buffer




def confirmar_salida_parqueadero(request, parqueadero_id):
    # Verifica si el usuario está autenticado
    if request.user.is_authenticated:
        try:
            # Obtiene el parqueadero por su ID
            parqueadero = Parqueadero.objects.get(id=parqueadero_id)

            # Verifica si el estado es 'Activo'
            if parqueadero.estado == 'Activo':
                fecha_salida = timezone.now()

                # Calcular horas estacionadas en Decimal
                horas_estacionado = (fecha_salida - parqueadero.fecha_ingreso).total_seconds() / 3600
                horas_estacionado_decimal = Decimal(horas_estacionado).quantize(Decimal('0.00'))  # Redondear a 2 decimales

                # Obtener el valor de fraccion_hora del conjunto relacionado (en Decimal)
                conjunto = Conjunto.objects.get(id=parqueadero.id_conjunto.id)
                valor_fraccion_hora = Decimal(conjunto.fraccion_hora)  # Asegúrate de que sea Decimal

                # Calcular el total a pagar
                valor_total = horas_estacionado_decimal * valor_fraccion_hora

                if valor_total < Decimal('700.00'):
                    valor_total = Decimal('700.00')


                # Generar el PDF de la factura
                pdf_buffer = generar_pdf_xhtml2pdf(parqueadero, valor_total, horas_estacionado, valor_fraccion_hora,fecha_salida)

                # Enviar el correo electrónico con el PDF adjunto
                try:
                    asunto = 'Confirmación de salida del parqueadero'
                    mensaje = (
                        f"Estimado usuario, su vehículo con placa {parqueadero.placa_vehiculo} ha salido del parqueadero.\n"
                        f"Total pagado: ${valor_total:,.2f}.\n"
                        f"Gracias por utilizar nuestro servicio. ¡Vuelva pronto!"
                    )
                    destinatario = parqueadero.contacto  # Asumimos que el contacto es un correo electrónico

                    email = EmailMessage(
                        asunto,
                        mensaje,
                        settings.EMAIL_HOST_USER,  # Remitente
                        [destinatario],  # Destinatario
                    )
                    email.attach('factura_parqueadero.pdf', pdf_buffer.getvalue(), 'application/pdf')
                    email.send(fail_silently=False)

                    # Si el correo se envía correctamente, actualiza el estado del parqueadero
                    parqueadero.total_calculado = valor_total
                    parqueadero.estado = 'Finalizado'
                    parqueadero.fecha_salida = fecha_salida
                    parqueadero.save()

                except Exception as e:
                    messages.error(request, f"Error al enviar el correo: {str(e)}")
                    # No actualizamos el estado si ocurre un error al enviar el correo
                    return redirect('/sistema/parqueadero/')

                # Si todo es exitoso, redirige al usuario
                return redirect('/sistema/parqueadero/')

            else:
                messages.error(request, "El parqueadero ya está finalizado.")
                return redirect("/sistema/parqueadero/")

        except Parqueadero.DoesNotExist:
            messages.error(request, "Parqueadero no encontrado.")
            return redirect("/sistema/parqueadero/")

    else:
        return redirect("/formularios/login/")















# citofonia 
from django.db.models import Q

def listar_citofonia(request):
    if request.user.is_authenticated:
        try:
            usuario_obj = Usuario.objects.get(email=request.user.email)
            if usuario_obj.rol.strip() in ['Administrador', 'Guarda']:
                query = request.GET.get('q', '')  # Obtener el término de búsqueda
                # Ordenar las propiedades por id (de menor a mayor)
                if query:
                    propiedades = Propiedad.objects.filter(numero__icontains=query).order_by('id')
                else:
                    propiedades = Propiedad.objects.all().order_by('id')
                
                return render(request, 'administra/citofonia/citofonia.html', {'Propiedades': propiedades, 'query': query})
            else:
                messages.error(request, "No tienes permiso para acceder a esta página.")
                return render(request, 'error.html', {"message": "Usted no tiene permisos para estar acá."})
        except Usuario.DoesNotExist:
            messages.error(request, "Usuario no encontrado.")
            return redirect("/sistema/login/")
    else:
        return redirect("/sistema/login/")

# gestion administrativa 

from django.core.paginator import Paginator
def listar_parqueaderos_gestion(request):
    if request.user.is_authenticated:
        try:
            usuario_obj = Usuario.objects.get(email=request.user.email)
            if usuario_obj.rol.strip() in ['Administrador', 'Guarda']:
                # Obtiene todos los parqueaderos con estado 'Finalizado' y los ordena por 'id' (de menor a mayor)
                parqueaderos = Parqueadero.objects.filter(estado='Finalizado').order_by('id')

                # Configura la paginación
                paginator = Paginator(parqueaderos, 15)  # 15 elementos por página
                page_number = request.GET.get('page')  # Número de página actual
                page_obj = paginator.get_page(page_number)

                return render(request, 'administra/parqueadero_admin/parqueadero.html', {'page_obj': page_obj})
            else:
                messages.error(request, "No tienes permiso para acceder a esta página.")
                return render(request, 'error.html', {"message": "Usted no tiene permisos para estar acá."})
        
        except Usuario.DoesNotExist:
            messages.error(request, "Usuario no encontrado.")
            return redirect("/sistema/login/")
    else:
        return redirect("/sistema/login/")
    

import openpyxl
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from .models import Parqueadero
from openpyxl.styles import Font
from openpyxl.styles import Font, Border, Side

def generar_informe_excel(request):
    if request.method == 'POST':
        # Obtener las fechas del formulario
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')

        # Convertir las fechas de formato string a datetime
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')  # Convertir a datetime
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')  # Convertir a datetime

        # Asegurarse de que la fecha de fin cubra hasta las 23:59:59
        fecha_fin = fecha_fin.replace(hour=23, minute=59, second=59, microsecond=999999)

        # Filtrar los parqueaderos según las fechas y el estado 'Finalizado', ordenados por id de menor a mayor
        parqueaderos = Parqueadero.objects.filter(
            fecha_ingreso__gte=fecha_inicio,
            fecha_ingreso__lte=fecha_fin,
            estado='Finalizado'
        ).order_by('id')

        # Crear un libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Informe Parqueaderos'

        # Escribir los encabezados
        ws['A1'] = 'Código'
        ws['B1'] = 'Ubicación'
        ws['C1'] = 'Placa Vehículo'
        ws['D1'] = 'Contacto'
        ws['E1'] = 'Casa'
        ws['F1'] = 'Fecha Ingreso'
        ws['G1'] = 'Fecha Salida'
        ws['H1'] = 'Estado'
        ws['I1'] = 'Guarda'
        ws['J1'] = 'Total CANCELADO'
        ws['K1'] = 'Total CANCELADO'
        # Crear un estilo de fuente en negrilla
        bold_font = Font(bold=True)

        # Crear un estilo de borde
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Lista de encabezados
        headers = ['Código', 'Ubicación', 'Placa Vehículo', 'Contacto', 'Casa', 'Fecha Ingreso', 
                'Fecha Salida', 'Estado', 'Guarda', 'Total CANCELADO', 'Total CANCELADO']

        # Aplicar los estilos a cada celda de encabezado
        for col, header in enumerate(headers, start=1):  # Enumerar desde la columna 1
            cell = ws.cell(row=1, column=col)  # Seleccionar la celda
            cell.value = header  # Asignar el valor del encabezado
            cell.font = bold_font  # Aplicar negrilla
            cell.border = thin_border  # Aplicar bordes


        # Escribir los datos de los parqueaderos
        row = 2  # Empezamos a escribir en la fila 2 (debajo de los encabezados)
        for parqueadero in parqueaderos:
            ws[f'A{row}'] = parqueadero.id
            ws[f'B{row}'] = parqueadero.tipo
            ws[f'C{row}'] = parqueadero.placa_vehiculo
            ws[f'D{row}'] = parqueadero.contacto
            ws[f'E{row}'] = parqueadero.casa
            ws[f'F{row}'] = parqueadero.fecha_ingreso.strftime('%Y-%m-%d %H:%M:%S')
            ws[f'G{row}'] = parqueadero.fecha_salida.strftime('%Y-%m-%d %H:%M:%S') if parqueadero.fecha_salida else ''
            ws[f'H{row}'] = parqueadero.estado
            ws[f'I{row}'] = f'{parqueadero.id_guarda.nombre} {parqueadero.id_guarda.apellido}'

            total_fila = row
            # Formatear el total calculado a dos decimales
            ws[f'J{row}'] = float(parqueadero.total_calculado) if parqueadero.total_calculado else 0
            ws[f'K2'] = f"=SUM(j2:j{total_fila})" 

            
            row += 1
            bold_font = Font(bold=True)
            ws[f'K2'].font = bold_font

        # Crear una respuesta HTTP para descargar el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Informe_Parqueaderos.xlsx'

        # Guardar el archivo en la respuesta HTTP
        wb.save(response)
        return response

    return render(request, 'ruta_del_template_con_formulario.html')



